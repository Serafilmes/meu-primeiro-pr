"""
relatorios_entrega.py — Camada 3 · Fatia 3: o DOCUMENTO FINAL da Entrega.

Gera um relatório CONGELADO da Planilha de Entrega — uma FOTO do momento, que
nunca mais muda (diferente do Google Sheets, que é sobrescrito o tempo todo).

Dois usos, MESMO motor:
  • RELATÓRIO DIÁRIO — o recorte de UM dia (todo fim de dia, evento ao vivo):
    "o que entrou/rolou hoje". Passa o bastão pro turno seguinte, vai pra produção.
  • RELATÓRIO FINAL  — o evento INTEIRO (uma vez, no fechamento): o arquivo
    definitivo de pesquisa/entrega contratual.

Cada relatório sai em DOIS formatos:
  • XLSX — a planilha de trabalho (capa + cabeçalho + autofiltro + linha congelada).
  • PDF  — o apresentável/assinável (paisagem, marca 6floor).

Os arquivos são salvos com CARIMBO DE HORA no nome (nunca sobrescrevem) em
projetos/<slug>/relatorios/ (ao lado do banco daquele projeto).

Fora do ciclo crítico: NUNCA toca mídia, só LÊ o banco. Offline. Sem IA.
Dependências: openpyxl (XLSX) e reportlab (PDF) — ambas grátis, puro-Python.
A fonte dos dados é banco_dados.montar_planilha_com_datas — a MESMA da tela
Entrega e do Sheets, então o relatório jamais diverge do que o operador vê.
"""

import os
import datetime

import banco_dados as bd


# ── PALETA 6floor (espelha marca/6floor_paleta.css) ─────────────────────────────
_TEAL       = "2BB58C"
_TEAL_FORTE = "1D9E75"
_BG_BASE    = "0B100E"
_BG_ELEV    = "131C19"
_TEXTO      = "1A2521"   # texto escuro sobre papel (o PDF/XLSX é claro, p/ impressão)
_LINHA_ALT  = "F2F7F5"   # zebra clara
_BORDA      = "C9D6D1"


# ── NOMES DE ARQUIVO ────────────────────────────────────────────────────────────

def _slug(texto):
    """Reduz um texto a um pedaço seguro pra nome de arquivo."""
    limpo = "".join(c if c.isalnum() else "_" for c in (texto or "").lower())
    while "__" in limpo:
        limpo = limpo.replace("__", "_")
    return limpo.strip("_") or "projeto"


def _base_nome(projeto, dia):
    """Prefixo do arquivo: entrega_<projeto>_<final|AAAAMMDD>_<AAAAMMDD_HHMM>."""
    agora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    recorte = "final" if not dia else dia.replace("-", "")
    return f"entrega_{_slug(projeto)}_{recorte}_{agora}"


# ── SELEÇÃO DAS LINHAS ──────────────────────────────────────────────────────────

def _colher(conn, dia):
    """Colunas + linhas do recorte pedido.

    dia = None  → tudo (relatório final).
    dia = 'AAAA-MM-DD' → só as linhas daquele dia (relatório diário).
    """
    colunas, linhas, datas = bd.montar_planilha_com_datas(conn)
    if dia:
        linhas = [ln for ln, d in zip(linhas, datas) if d == dia]
    return colunas, linhas


# ── XLSX (a planilha de trabalho) ───────────────────────────────────────────────

def _gerar_xlsx(caminho, titulo, subtitulo, colunas, linhas, gerado_por):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Entrega"

    rotulos = [c["rotulo"] for c in colunas]
    n_col = max(1, len(rotulos))

    # ── Capa (3 linhas de contexto acima da tabela) ──
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=16, color=_TEAL_FORTE)
    ws.cell(row=2, column=1, value=subtitulo).font = Font(size=11, color="555555")
    carimbo = (f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
               + (f" por {gerado_por}" if gerado_por else "")
               + f"  ·  {len(linhas)} take(s)")
    ws.cell(row=3, column=1, value=carimbo).font = Font(size=10, italic=True, color="777777")

    LINHA_CAB = 5  # o cabeçalho da tabela começa aqui

    # ── Cabeçalho ──
    fill_cab = PatternFill("solid", fgColor=_BG_ELEV)
    fonte_cab = Font(bold=True, color="FFFFFF", size=10)
    borda_fina = Border(*[Side(style="thin", color=_BORDA)] * 4)
    for j, rot in enumerate(rotulos, start=1):
        cel = ws.cell(row=LINHA_CAB, column=j, value=rot)
        cel.fill = fill_cab
        cel.font = fonte_cab
        cel.alignment = Alignment(vertical="center", wrap_text=True)
        cel.border = borda_fina

    # ── Corpo (com zebra) ──
    fill_zebra = PatternFill("solid", fgColor=_LINHA_ALT)
    for i, linha in enumerate(linhas):
        r = LINHA_CAB + 1 + i
        for j in range(n_col):
            val = linha[j] if j < len(linha) else ""
            cel = ws.cell(row=r, column=j + 1, value=val)
            cel.font = Font(size=10, color=_TEXTO)
            cel.alignment = Alignment(vertical="center", wrap_text=False)
            cel.border = borda_fina
            if i % 2 == 1:
                cel.fill = fill_zebra

    # ── Largura das colunas (pelo maior conteúdo, com teto) ──
    for j in range(n_col):
        maior = len(rotulos[j]) if j < len(rotulos) else 8
        for linha in linhas:
            if j < len(linha):
                maior = max(maior, len(str(linha[j])))
        ws.column_dimensions[get_column_letter(j + 1)].width = min(max(maior + 2, 10), 48)

    # ── Autofiltro + painel congelado (cabeçalho fixa ao rolar) ──
    ultima_col = get_column_letter(n_col)
    ws.auto_filter.ref = f"A{LINHA_CAB}:{ultima_col}{LINHA_CAB + len(linhas)}"
    ws.freeze_panes = f"A{LINHA_CAB + 1}"

    wb.save(caminho)
    return caminho


# ── PDF (o apresentável/assinável) ──────────────────────────────────────────────

def _gerar_pdf(caminho, titulo, subtitulo, colunas, linhas, gerado_por):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(
        caminho, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=titulo,
    )
    estilos = getSampleStyleSheet()
    st_titulo = ParagraphStyle("t", parent=estilos["Title"], fontSize=18,
                               textColor=colors.HexColor("#" + _TEAL_FORTE), spaceAfter=2)
    st_sub = ParagraphStyle("s", parent=estilos["Normal"], fontSize=10,
                            textColor=colors.HexColor("#555555"))
    st_carimbo = ParagraphStyle("c", parent=estilos["Normal"], fontSize=8,
                                textColor=colors.HexColor("#888888"))
    st_cel = ParagraphStyle("cel", parent=estilos["Normal"], fontSize=6.5, leading=8)
    st_cab = ParagraphStyle("cab", parent=estilos["Normal"], fontSize=7,
                            leading=8, textColor=colors.white, fontName="Helvetica-Bold")

    elementos = [
        Paragraph("∞ 6floor", ParagraphStyle("marca", parent=estilos["Normal"],
                  fontSize=11, textColor=colors.HexColor("#" + _TEAL))),
        Paragraph(titulo, st_titulo),
        Paragraph(subtitulo, st_sub),
        Paragraph(f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
                  + (f" por {gerado_por}" if gerado_por else "")
                  + f"  ·  {len(linhas)} take(s)", st_carimbo),
        Spacer(1, 6 * mm),
    ]

    rotulos = [c["rotulo"] for c in colunas]
    dados = [[Paragraph(str(r), st_cab) for r in rotulos]]
    for linha in linhas:
        dados.append([Paragraph(str(v).replace("\n", " "), st_cel) for v in linha])

    if len(dados) == 1:   # só o cabeçalho — nenhuma linha no recorte
        dados.append([Paragraph("(sem takes neste recorte)", st_cel)]
                     + [Paragraph("", st_cel)] * (len(rotulos) - 1))

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + _BG_ELEV)),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#" + _BORDA)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#" + _LINHA_ALT)]),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    elementos.append(tabela)
    doc.build(elementos)
    return caminho


# ── PONTO DE ENTRADA ────────────────────────────────────────────────────────────

def gerar(conn, projeto, pasta_saida, dia=None, gerado_por=""):
    """Gera o relatório congelado (XLSX + PDF) do recorte pedido.

    Args:
        conn        — conexão aberta do banco do projeto ativo.
        projeto     — nome legível do projeto (vai na capa e no nome do arquivo).
        pasta_saida — pasta onde salvar (ex.: projetos/<slug>/relatorios).
        dia         — None = relatório FINAL (tudo); 'AAAA-MM-DD' = DIÁRIO.
        gerado_por  — nome do operador logado (vai na capa; opcional).

    Returns:
        dict {ok, xlsx, pdf, n_linhas, titulo, dia} — caminhos absolutos dos
        dois arquivos gerados.
    """
    os.makedirs(pasta_saida, exist_ok=True)
    colunas, linhas = _colher(conn, dia)

    if not colunas:
        # Nenhuma coluna visível no molde — o relatório sairia vazio de conteúdo.
        # Erro claro (a rota do Flask mostra a mensagem) em vez de um arquivo inútil.
        raise ValueError(
            "Nenhuma coluna visível na Entrega — ative colunas em "
            "“⚙ Configurar colunas” (/molde) antes de gerar o relatório."
        )

    if dia:
        try:
            legivel = datetime.datetime.strptime(dia, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            legivel = dia
        titulo = "Relatório de Entrega — Diário"
        subtitulo = f"{projeto}  ·  dia {legivel}"
    else:
        titulo = "Relatório de Entrega — Final"
        subtitulo = f"{projeto}  ·  evento completo"

    base = os.path.join(pasta_saida, _base_nome(projeto, dia))
    xlsx = _gerar_xlsx(base + ".xlsx", titulo, subtitulo, colunas, linhas, gerado_por)
    pdf = _gerar_pdf(base + ".pdf", titulo, subtitulo, colunas, linhas, gerado_por)

    return {
        "ok": True,
        "xlsx": os.path.abspath(xlsx),
        "pdf": os.path.abspath(pdf),
        "n_linhas": len(linhas),
        "titulo": titulo,
        "dia": dia,
    }
